import streamlit as st
import math
import pandas as pd
import matplotlib.pyplot as plt
import matplotlib
matplotlib.use("Agg")
from io import BytesIO
from openpyxl.styles import PatternFill

st.set_page_config(page_title="Packing Distribution Solver", layout="wide")

st.markdown("""
<style>
body { background-color: #bd1515; }
h1, h2, h3 { color: #E60012; font-weight: 700; }
.toyota-card {
  background: white;
  border-radius: 16px;
  padding: 18px;
  border-left: 6px solid #E60012;
  box-shadow: 0 4px 14px rgba(0,0,0,0.08);
  margin-bottom: 12px;
}
button { background-color: #E60012 !important; color: white !important; border-radius: 10px !important; }
</style>

<div class='toyota-card'>
<h1>üè≠ Packing Solver</h1>
</div>
""", unsafe_allow_html=True)

if "prev_current_packing" not in st.session_state:
    st.session_state.prev_current_packing = None

st.markdown("## Bulk Copy Input (Paste from Raw Data)")

# Input
NUM_PARTS = st.number_input("Jumlah Part", min_value=1, step=1)
NUM_DAYS = st.number_input("Jumlah Hari", min_value=31, step=1)

part_text = st.text_area("No Part (Vertical)", height=150)
flag_text = st.text_area("Flags X/O (Vertical)", height=150)
lot_text = st.text_area("Lot Size (Vertical)", height=150)

pattern_text = st.text_area("Packing Pattern (Horizontal)")

origin_text = st.text_area(
    "Packing Pattern Original (PIECES Level ‚Äî Rows = Part, Columns = Days)", 
    height=200
)

TOP_N_ADJUST = st.number_input(
    "Jumlah Part Flag O yang akan Disesuaikan (Top Largest)",
    min_value=0,
    step=1,
    value=10
)

st.markdown("## MSP Uploaded Format Bulk Input")

Dock_text = st.text_area("Dock Code (Vertical)", height=150)
FirmMonth_single = st.text_input("Firm Packing Month (Input Sekali)")
CarFamily_single = st.text_input("Car Family Code (Input Sekali)")
ReExport_text = st.text_area("Re-Export Code (Vertical)", height=150)
Kanban_text = st.text_area("Kanban No (Vertical)", height=150)
AICOCEPT_text = st.text_area("AICO/CEPT_N (Vertical)", height=150)
Series_text = st.text_area("Series (Vertical)", height=150)
PartNameManual_text = st.text_area("Part Name (Vertical)", height=150)

output_filename = st.text_input("Nama File Output Excel (Akhiri .xlsx)", value="Output Solver.xlsx")

run_solver = st.button("Run Solver")

if run_solver:

    errors = []

    PartNames = [x.strip() for x in part_text.splitlines() if x.strip()]
    flags = [x.strip() for x in flag_text.splitlines() if x.strip()]
    lot_lines = [x.strip() for x in lot_text.splitlines() if x.strip()]
    pattern_vals = pattern_text.split()
    origin_lines = [row for row in origin_text.splitlines() if row.strip()]

    DockCode = [x.strip() for x in Dock_text.splitlines() if x.strip()]
    ReExportCode = [x.strip() for x in ReExport_text.splitlines() if x.strip()]
    KanbanNo = [x.strip() for x in Kanban_text.splitlines() if x.strip()]
    AICOCEPT = [x.strip() for x in AICOCEPT_text.splitlines() if x.strip()]
    SeriesCode = [x.strip() for x in Series_text.splitlines() if x.strip()]
    PartNameManual = [x.strip() for x in PartNameManual_text.splitlines() if x.strip()]

    # Message Error
    if len(PartNames) != NUM_PARTS:
        errors.append("‚ùå No Part: Jumlah Input Tidak Sesuai")

    if len(flags) != NUM_PARTS:
        errors.append("‚ùå Flag Part: Jumlah Input Tidak Sesuai")
    else:
        invalid_flags = [f for f in flags if f.upper() not in ["X", "O"]]
        if invalid_flags:
            errors.append("‚ùå Flag Part: hanya boleh X atau O")

    if len(lot_lines) != NUM_PARTS:
        errors.append("‚ùå Lot Size: Jumlah Input Tidak Sesuai")

    if len(pattern_vals) != NUM_DAYS:
        errors.append("‚ùå Packing Pattern: Jumlah Input Tidak Sesuai")

    if len(origin_lines) != NUM_PARTS:
        errors.append("‚ùå Packing Pattern Original: Jumlah Input Tidak Sesuai")
    else:
        for row in origin_lines:
            if len(row.split()) != NUM_DAYS:
                errors.append("‚ùå Packing Pattern Original: Jumlah Input Tidak Sesuai")
                break

    if len(DockCode) != NUM_PARTS:
        errors.append("‚ùå Dock Code: Jumlah Input Tidak Sesuai")
    if len(ReExportCode) != NUM_PARTS:
        errors.append("‚ùå Re-Export Code: Jumlah Input Tidak Sesuai")
    if len(KanbanNo) != NUM_PARTS:
        errors.append("‚ùå Kanban No: Jumlah Input Tidak Sesuai")
    if len(AICOCEPT) != NUM_PARTS:
        errors.append("‚ùå AICO/CEPT_N: Jumlah Input Tidak Sesuai")
    if len(SeriesCode) != NUM_PARTS:
        errors.append("‚ùå Series: Jumlah Input Tidak Sesuai")
    if len(PartNameManual) != NUM_PARTS:
        errors.append("‚ùå Part Name: Jumlah Input Tidak Sesuai")

    if errors:
        for e in errors:
            st.error(e)
        st.stop()

    # Original Solver
    LotSize = [float(x.strip()) for x in lot_text.splitlines() if x.strip()]
    PackingPattern = list(map(float, pattern_text.split()))
    IsWorkingDay = [1 if p != 0 else 0 for p in PackingPattern]
    origin_pieces = [list(map(int, row.split())) for row in origin_text.splitlines() if row.strip()]

    origin = []
    for i in range(len(origin_pieces)):
        row_box = [origin_pieces[i][d] / LotSize[i] for d in range(len(origin_pieces[i]))]
        origin.append([round(x) for x in row_box])

    OriginalPacking = [0] * len(PackingPattern)
    for i in range(len(origin)):
        for d in range(len(origin[i])):
            if IsWorkingDay[d] == 1:
                OriginalPacking[d] += origin[i][d]

    TotalPart = [sum(origin[i]) for i in range(len(origin))]

    NUM_PARTS = len(PartNames)
    NUM_DAYS = len(PackingPattern)

    total_current = sum(TotalPart)
    total_pattern = sum(PackingPattern)
    X = total_current / total_pattern if total_pattern > 0 else 0
    TargetPacking = [p * X for p in PackingPattern]

    result = [[0]*NUM_DAYS for _ in range(NUM_PARTS)]
    CurrentPacking = [0]*NUM_DAYS
    workdays = [d for d in range(NUM_DAYS) if IsWorkingDay[d] == 1]

    for i in range(NUM_PARTS):
        if flags[i].upper() == "X":
            for d in range(NUM_DAYS):
                if IsWorkingDay[d] == 1:
                    result[i][d] = origin[i][d]
                    CurrentPacking[d] += origin[i][d]

    parts_O_all = [i for i in range(NUM_PARTS) if flags[i].upper() == "O"]
    parts_O_sorted = sorted(parts_O_all, key=lambda i: TotalPart[i], reverse=True)

    parts_O_adjust = parts_O_sorted[:TOP_N_ADJUST]
    parts_O_fixed = parts_O_sorted[TOP_N_ADJUST:]

    for i in parts_O_fixed:
        for d in range(NUM_DAYS):
            if IsWorkingDay[d] == 1:
                result[i][d] = origin[i][d]
                CurrentPacking[d] += origin[i][d]

    for i in parts_O_adjust:
        remaining = TotalPart[i]
        while remaining > 0:
            capacities = [
                max(0, TargetPacking[d] + 15 - CurrentPacking[d]) if d in workdays else 0
                for d in range(NUM_DAYS)
            ]
            total_capacity = sum(capacities)
            if total_capacity == 0:
                break

            temp_assign = [0]*NUM_DAYS
            for d in range(NUM_DAYS):
                if capacities[d] > 0:
                    temp_assign[d] = min(math.floor(capacities[d]/total_capacity*remaining), remaining)

            leftover = remaining - sum(temp_assign)

            if leftover > 0:
                for d in sorted(range(NUM_DAYS), key=lambda x: capacities[x], reverse=True):
                    if leftover <= 0:
                        break
                    if capacities[d] - temp_assign[d] > 0:
                        temp_assign[d] += 1
                        leftover -= 1

            for d in range(NUM_DAYS):
                result[i][d] += temp_assign[d]
                CurrentPacking[d] += temp_assign[d]

            remaining = 0

    # Dashboard
    st.markdown(f"## üìä Production Dashboard (CFC: **{CarFamily_single}**)")

    col1, col2, col3, col4 = st.columns(4)
    non_zero_current = [x for x in CurrentPacking if x > 0]
    col1.metric("Total Units", sum(TotalPart))
    col2.metric("Max Target", max(TargetPacking))
    col3.metric("Max Current", max(CurrentPacking))
    col4.metric("Min Current (0 not included)", min(non_zero_current) if non_zero_current else 0)

    col_left, col_right = st.columns(2)

    with col_left:
        fig1, ax1 = plt.subplots(figsize=(6, 3))
        ax1.bar(range(NUM_DAYS), OriginalPacking, alpha=0.45)
        ax1.plot(TargetPacking, marker="o")
        ax1.set_title("Before", fontsize=9)
        ax1.tick_params(labelsize=7)
        st.pyplot(fig1)

    with col_right:
        fig2, ax2 = plt.subplots(figsize=(6, 3))
        ax2.bar(range(NUM_DAYS), CurrentPacking, alpha=0.45)
        ax2.plot(TargetPacking, marker="o")
        ax2.set_title("After", fontsize=9)
        ax2.tick_params(labelsize=7)
        st.pyplot(fig2)

    # Output Excel
    df_pieces = pd.DataFrame(result, columns=[f"Hari_{i+1}" for i in range(NUM_DAYS)])
    df_pieces.insert(0, "Part", PartNames)
    df_pieces["Total Part (Box)"] = [sum(row) for row in result]

    df_pieces.loc["TOTAL"] = ["TOTAL"] + CurrentPacking + [sum(CurrentPacking)]
    df_pieces.loc["TARGET"] = ["TARGET"] + [round(x, 1) for x in TargetPacking] + [""]

    result_box = [[result[i][d] * LotSize[i] for d in range(NUM_DAYS)] for i in range(NUM_PARTS)]
    df_box = pd.DataFrame(result_box, columns=[f"Hari_{i+1}" for i in range(NUM_DAYS)])
    df_box.insert(0, "Part", PartNames)

    rows_O = []
    names_O = []
    for i in parts_O_adjust:
        rows_O.append(result_box[i])
        names_O.append(PartNames[i])

    df_box_O = pd.DataFrame(rows_O, columns=[f"Hari_{i+1}" for i in range(NUM_DAYS)])
    df_box_O.insert(0, "Part", names_O)

    msp_rows_minus = []
    msp_rows_plus = []

    for i in parts_O_adjust:

        base_meta = {
            "MD": "U",
            "FT": "M",
            "Company Code / Importer": "807D",
            "Receiving Plant Code": "4",
            "Dock Code": DockCode[i],
            "Supplier Code / Exporter": " ",
            "Supplier Plant Code": "",
            "Shipping Dock": "",
            "Cross Dock": "",
            "Cross Dock Plant Code": "",
            "MSP Order Type": "R",
            "Firm Packing Month (N Month)": FirmMonth_single,
            "Car Family Code": CarFamily_single,
            "Re-Export Code": ReExportCode[i],
            "Part No": PartNames[i],
            "Order Lot Size": LotSize[i],
            "Kanban No": KanbanNo[i],
        }

        row_minus = base_meta.copy()
        row_minus["Sign"] = "-"
        for d in range(NUM_DAYS):
            row_minus[f"N-{d+1}"] = origin_pieces[i][d]

        row_minus["Source Code"] = "4"
        row_minus["AICO/CEPT_N"] = AICOCEPT[i]
        row_minus["Series"] = SeriesCode[i]
        row_minus["Life Cycle Code"] = "0"
        row_minus["Supplier / Exporter Name"] = "TOYOTA MOTOR THAILAND"
        row_minus["Part Name"] = PartNameManual[i]

        row_plus = base_meta.copy()
        row_plus["Sign"] = "+"
        for d in range(NUM_DAYS):
            row_plus[f"N-{d+1}"] = result_box[i][d]

        row_plus["Source Code"] = "4"
        row_plus["AICO/CEPT_N"] = AICOCEPT[i]
        row_plus["Series"] = SeriesCode[i]
        row_plus["Life Cycle Code"] = "0"
        row_plus["Supplier / Exporter Name"] = "TOYOTA MOTOR THAILAND"
        row_plus["Part Name"] = PartNameManual[i]

        msp_rows_minus.append(row_minus)
        msp_rows_plus.append(row_plus)

    msp_rows = msp_rows_minus + msp_rows_plus
    df_msp = pd.DataFrame(msp_rows)

    # Export
    output = BytesIO()
    with pd.ExcelWriter(output, engine="openpyxl") as writer:
        df_pieces.to_excel(writer, sheet_name="Box Level", index=False)
        df_box.to_excel(writer, sheet_name="Pieces Level", index=False)
        df_box_O.to_excel(writer, sheet_name="Top N Pieces Level", index=False)
        df_msp.to_excel(writer, sheet_name="MSP Format", index=False)

        workbook = writer.book
        ws = writer.sheets["MSP Format"]

        yellow_fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")

        sign_col_index = list(df_msp.columns).index("Sign") + 1

        for row_idx in range(2, len(df_msp) + 2):
            sign_val = ws.cell(row=row_idx, column=sign_col_index).value
            if sign_val == "-":
                for col_idx in range(1, len(df_msp.columns) + 1):
                    ws.cell(row=row_idx, column=col_idx).fill = yellow_fill

    st.download_button(
        label="üì• Download Excel Solver Output",
        data=output.getvalue(),
        file_name=output_filename,
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )

    st.success("Solver selesai")

